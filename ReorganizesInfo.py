'''
Take informations from auction.xlsx (which is in several tabs) and reorganizes
it in a single sheet in auctionresults.xlsx, so i can use formules
to calculate profits and etc in the same sheet.

Probably possible to just do it in auction.xlsx from the start, but i didnt knew how.

'''
import openpyxl

wb1 = load_workbook('auction.xlsx')
wb2 = load_workbook('auctionresults.xlsx')

a2 = wb2.active
my_sheets = wb1.sheetnames


add = -8


for sheets in my_sheets:
    sheet1 = wb1[f'{sheets}']
    add = add + 9
    maxr = sheet1.max_row
    maxc = sheet1.max_column
    for r in range (1, maxr + 1):
        for c in range (1, maxc + 1):
            a2.cell(row=r+add,column=c).value = sheet1.cell(row=r,column=c).value
            wb2.save('auctionresults.xlsx')
            
pyautogui.moveTo(0, 0)
