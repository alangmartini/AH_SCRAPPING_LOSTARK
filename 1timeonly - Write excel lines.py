from openpyxl import load_workbook

# --------------------------------------------
'''
Made e little program for inserting the header titles in excel, just to make it quick.
Only necessary when creating a new auctionresults file.


'''


wb1 = load_workbook('auction.xlsx')
wb2 = load_workbook('auctionresults.xlsx')

my_sheets = wb1.sheetnames
a2 = wb2.active
add = -8


time.sleep(1)
title = 1
for sheets in my_sheets:
    a2.cell(row=title,column=1).value = sheets
    title = title + 9
    wb2.save('auctionresults.xlsx')
pyautogui.moveTo(0,0) 
